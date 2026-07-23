#!/usr/bin/env python3
"""Build the approved-release index from MBK's publishing workbook.

This is an offline administrative import.  It does not change the workbook.
Only rows with a title and a recognized live publisher URL are accepted.
"""

from __future__ import annotations

import argparse
import gzip
import json
import os
import re
import sys
from collections import Counter
from datetime import date, datetime
from urllib.parse import urlparse

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from exemplar_corpus import (  # noqa: E402
    _PLATFORM_HOSTS,
    _tokens,
    infer_intents,
    infer_vertical,
    normalize_platform,
)


def _iso_date(value) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def _title_pattern(title: str) -> str:
    pattern = re.sub(r"\b(?:19|20)\d{2}\b", "[YEAR]", title)
    pattern = re.sub(r"\$\d+(?:[.,]\d+)*", "[PRICE]", pattern)
    pattern = re.sub(r"\b\d+(?:\.\d+)?%?\b", "[NUMBER]", pattern)
    parts = re.split(r"(\s[:—–-]\s|:\s)", pattern, maxsplit=1)
    if len(parts) > 2:
        lead = parts[0]
        lead = re.sub(
            r"^.{2,80}?(?=\s(?:Review|Reviews|Reviewed|Update|Analysis)\b)",
            "[PRODUCT]",
            lead,
            flags=re.I,
        )
        return lead + "".join(parts[1:])
    return re.sub(
        r"^.{2,80}?\b(Review|Reviews|Reviewed)\b",
        r"[PRODUCT] \1",
        pattern,
        flags=re.I,
    )


def _product_name_from_title(title: str) -> str:
    """Extract a conservative product/offering candidate from a release title."""
    clean = re.sub(r"\s+", " ", title).strip()
    separators = (
        r"\s+(?:Review|Reviews|Reviewed|Complaints|Ingredients|Side Effects|"
        r"Pricing|Price|Benefits|Results|Official Website|Update)\b",
        r"\s*[:|—–]\s*",
        r"\s+-\s+",
    )
    candidates = [clean]
    for pattern in separators:
        match = re.search(pattern, clean, flags=re.I)
        if match and match.start() >= 2:
            candidates.append(clean[:match.start()].strip(" :-|—–"))
    candidate = min(candidates, key=len)
    candidate = re.sub(
        r"^(?:New|Updated|Breaking|Exclusive)\s+", "", candidate, flags=re.I
    )
    if len(candidate) < 2 or len(candidate) > 120:
        return clean[:120].strip()
    return candidate


def _entity_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _source_domain(url: str) -> str:
    return urlparse(url or "").netloc.lower().removeprefix("www.")


def _header_map(row) -> dict[str, int]:
    result = {}
    for idx, value in enumerate(row):
        key = re.sub(r"\s+", " ", str(value or "").strip().upper())
        if key:
            result.setdefault(key, idx)
    return result


def _recognized_live_url(row, live_indexes) -> str:
    allowed = {host for hosts in _PLATFORM_HOSTS.values() for host in hosts}
    for idx in live_indexes:
        if idx >= len(row):
            continue
        value = str(row[idx] or "").strip()
        host = urlparse(value).netloc.lower().removeprefix("www.")
        if value.startswith("http") and host in allowed:
            return value
    return ""


def build(workbook_path: str) -> dict:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    releases = {}
    rejected = Counter()
    sheets = Counter()

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        first = next(rows, ())
        headers = _header_map(first)

        # The dedicated NW sheet intentionally has no header row.
        if sheet.title == "NW" and "TITLE" not in headers:
            headers = {
                "DATE": 0, "TYPE": 2, "PR SITE": 3, "TITLE": 4,
                "SOURCE": 5, "PREVIEW URL": 6, "PRETTY LINKS": 7,
                "LIVE LINK URL": 13,
            }
            rows = sheet.iter_rows(min_row=2, values_only=True)

        title_idx = headers.get("TITLE")
        if title_idx is None:
            continue
        live_indexes = [
            idx for name, idx in headers.items()
            if "LIVE" in name and "URL" in name
        ]
        if not live_indexes:
            continue

        def idx(*names):
            return next((headers[name] for name in names if name in headers), None)

        date_idx = idx("DATE", "2.0")
        type_idx = idx("TYPE")
        platform_idx = idx("PR SITE", "PR")
        source_idx = idx("SOURCE")
        preview_idx = idx("PREVIEW URL")
        pretty_idx = idx("PRETTY LINKS")

        for row_number, row in enumerate(rows, 2):
            def value(column):
                return row[column] if column is not None and column < len(row) else None

            title = str(value(title_idx) or "").strip()
            if not title or title.startswith(("↓", "↑", "➜")):
                rejected["non_release_title"] += 1
                continue
            live_url = _recognized_live_url(row, live_indexes)
            if not live_url:
                rejected["no_recognized_live_url"] += 1
                continue

            platform_raw = str(value(platform_idx) or "").strip()
            platform = normalize_platform(platform_raw, live_url)
            source_url = str(value(source_idx) or "").strip()
            preview_url = str(value(preview_idx) or "").strip()
            pretty_url = str(value(pretty_idx) or "").strip()
            release_type = str(value(type_idx) or "").strip()
            published_date = _iso_date(value(date_idx))
            vertical = infer_vertical(title, source_url, release_type)

            record = {
                "title": title,
                "title_pattern": _title_pattern(title),
                "platform": platform,
                "vertical": vertical,
                "intents": infer_intents(title),
                "tokens": sorted(_tokens(title + " " + source_url)),
                "published_date": published_date,
                "release_type": release_type,
                "source_url": source_url if source_url.startswith("http") else "",
                "preview_url": preview_url if preview_url.startswith("http") else "",
                "pretty_url": pretty_url if pretty_url.startswith("http") else "",
                "live_url": live_url,
                "workbook_sheet": sheet.title,
                "workbook_row": row_number,
            }
            if published_date:
                try:
                    year = int(published_date[:4])
                    record["recency_score"] = max(0, min((year - 2023) * 0.1, 0.3))
                except ValueError:
                    record["recency_score"] = 0
            else:
                record["recency_score"] = 0

            existing = releases.get(live_url)
            if existing is None or len(record["source_url"]) > len(existing["source_url"]):
                releases[live_url] = record
            sheets[sheet.title] += 1

    ordered = sorted(
        releases.values(),
        key=lambda item: (
            item["platform"], item["vertical"], item["published_date"], item["title"]
        ),
    )
    entity_map = {}
    for release in ordered:
        product = _product_name_from_title(release["title"])
        key = _entity_key(product)
        if not key:
            continue
        entity = entity_map.setdefault(key, {
            "product_or_offering": product,
            "brand_candidate": product,
            "company_status": "DOMAIN CANDIDATE — requires source verification",
            "source_domains": set(),
            "source_urls": set(),
            "platforms": set(),
            "verticals": set(),
            "intents": set(),
            "live_release_urls": set(),
            "release_count": 0,
            "first_seen": "",
            "last_seen": "",
        })
        domain = _source_domain(release.get("source_url", ""))
        if domain:
            entity["source_domains"].add(domain)
        if release.get("source_url"):
            entity["source_urls"].add(release["source_url"])
        entity["platforms"].add(release["platform"])
        entity["verticals"].add(release["vertical"])
        entity["intents"].update(release["intents"])
        entity["live_release_urls"].add(release["live_url"])
        entity["release_count"] += 1
        seen = release.get("published_date", "")
        if seen:
            if not entity["first_seen"] or seen < entity["first_seen"]:
                entity["first_seen"] = seen
            if not entity["last_seen"] or seen > entity["last_seen"]:
                entity["last_seen"] = seen

    entities = []
    for entity in entity_map.values():
        for field in (
            "source_domains", "source_urls", "platforms", "verticals",
            "intents", "live_release_urls",
        ):
            entity[field] = sorted(entity[field])
        entities.append(entity)
    entities.sort(key=lambda item: (-item["release_count"], item["product_or_offering"]))

    return {
        "_meta": {
            "description": "Published MBK releases for structure-only exemplar retrieval",
            "source_workbook": os.path.basename(workbook_path),
            "built_at": datetime.now().isoformat(timespec="seconds"),
            "approval_rule": "Title + recognized live publisher URL",
            "release_count": len(ordered),
            "unique_product_or_offering_count": len(entities),
            "platform_counts": dict(Counter(x["platform"] for x in ordered)),
            "vertical_counts": dict(Counter(x["vertical"] for x in ordered)),
            "sheet_counts": dict(sheets),
            "rejected": dict(rejected),
        },
        "releases": ordered,
        "entity_inventory": entities,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument(
        "--output",
        default=os.path.join(ROOT, "approved_release_index.json.gz"),
    )
    args = parser.parse_args()
    payload = build(os.path.abspath(args.workbook))
    with gzip.open(args.output, "wt", encoding="utf-8", compresslevel=9) as handle:
        json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))
    print(json.dumps(payload["_meta"], indent=2))
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
